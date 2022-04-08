from boto3 import client
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

try:
    import IhiwRestAccess
    import ParseExcel
    import ParseXml
    import Validation
    import S3_Access
    import ImmunogenicEpitopesValidator
except Exception as e:
    print('Failed in importing files: ' + str(e))
    from Common import IhiwRestAccess
    from Common import ParseExcel
    from Common import ParseXml
    from Common import Validation
    from Common import S3_Access
    import ImmunogenicEpitopesValidator

s3 = client('s3')
from sys import exc_info

import zipfile
import io

def createUploadEntriesForReport(summaryFileName=None, zipFileName=None):
    # TODO: This should be a standalone upload, not a child upload. Need some work on this part.
    # Probably need to modify the REST method to allow validation user to do this.

    # TODO: This will also make multiple copies. I should check if the report file already exists and/or (probably) overwrite it
    parentUploadName = '1497_1615205312528_PROJECT_DATA_MATRIX_ProjectReport'
    url = IhiwRestAccess.getUrl()
    token = IhiwRestAccess.getToken(url=url)

    if(url is not None and token is not None and len(url)>0 and len(token)>0):

        IhiwRestAccess.createConvertedUploadObject(newUploadFileName=summaryFileName
                                                   , newUploadFileType='OTHER'
                                                   , previousUploadFileName=parentUploadName
                                                   , url=url, token=token)
        IhiwRestAccess.createConvertedUploadObject(newUploadFileName=zipFileName
                                                   , newUploadFileType='OTHER'
                                                   , previousUploadFileName=parentUploadName
                                                   , url=url, token=token)

        IhiwRestAccess.setValidationStatus(uploadFileName=parentUploadName, isValid=True,
                                           validationFeedback='Valid Report.', validatorType='PROJECT_REPORT', url=url,
                                           token=token)
        IhiwRestAccess.setValidationStatus(uploadFileName=summaryFileName, isValid=True,
                                           validationFeedback='Valid Report.', validatorType='PROJECT_REPORT', url=url,
                                           token=token)
        IhiwRestAccess.setValidationStatus(uploadFileName=zipFileName, isValid=True,
                                           validationFeedback='Valid Report.', validatorType='PROJECT_REPORT', url=url,
                                           token=token)
    else:
        raise Exception('Could not create login token when creating upload entries for report files.')

def getTransplantationReportSpreadsheet(donorTyping=None, recipientTyping=None, recipHamlPreTxFilenames=None, recipHamlPostTxFilenames=None, s3=None, bucket=None, transplantationIndex=None, recipientSampleId=None):
    recipPreTxAntibodyData = ParseXml.parseHamlFileForBeadData(hamlFileNames=recipHamlPreTxFilenames, s3=s3, bucket=bucket, sampleIdQuery=recipientSampleId)
    recipPostTxAntibodyData = ParseXml.parseHamlFileForBeadData(hamlFileNames=recipHamlPostTxFilenames, s3=s3, bucket=bucket, sampleIdQuery=recipientSampleId)
    transplantationReportSpreadsheet = ParseExcel.createExcelTransplantationReport(donorTyping=donorTyping, recipientTyping=recipientTyping, recipPreTxAntibodyData=recipPreTxAntibodyData, recipPostTxAntibodyData=recipPostTxAntibodyData, preTxFileNames=recipHamlPreTxFilenames, postTxFileNames=recipHamlPostTxFilenames, transplantationIndex=transplantationIndex)
    return transplantationReportSpreadsheet, recipPreTxAntibodyData, recipPostTxAntibodyData

def getDataMatrixValue(validatedWorkbook=None, columnName=None, currentExcelRow=None, firstSheet=None):
    try:
        data = firstSheet[validatedWorkbook.columnNameLookup[columnName] + str(currentExcelRow)].value
        if data is None:
            return '?'
        else:
            return data
    except Exception as e:
        return '?'

def parseGlStrings(glstrings=None):
    # Parse GL String, locus delimiters. Keep copies of genes together, I guess.
    #print('parsing GL String:' + str(glstring))
    glStringTyping={}

    # Split by locus
    for glstring in glstrings:
        for locusToken in glstring.split('^'):
            #print('Found locus:' + str(locusToken))
            if('A*' in locusToken and 'MICA*' not in locusToken):
                #glStringTyping['A']=locusToken
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'A':locusToken})
            elif ('B*' in locusToken and 'MICB*' not in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'B':locusToken})
            elif ('C*' in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'C':locusToken})
            elif ('DRB1*' in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'DRB1':locusToken})
            elif ('DRB3*' in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'DRB3':locusToken})
            elif ('DRB4*' in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'DRB4':locusToken})
            elif ('DRB5*' in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'DRB5':locusToken})
            elif ('DQB1*' in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'DQB1':locusToken})
            elif ('DQA1*' in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'DQA1':locusToken})
            elif ('DPB1*' in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'DPB1':locusToken})
            elif ('DPA1*' in locusToken):
                glStringTyping = updateTypings(typings=glStringTyping, newTypings={'DPA1':locusToken})
            elif('HLA-E*' in locusToken or 'HLA-F*' in locusToken or 'HLA-G*' in locusToken or 'HLA-H*' in locusToken or 'MICA*' in locusToken or 'MICB*' in locusToken):
                # Not a problem but we're not using these.
                pass
            else:
                print('Unknown Locus:' + str(locusToken))

    return glStringTyping

def updateTypings(typings=None, newTypings=None):
    for locus in newTypings.keys():
        if(locus not in typings.keys() or typings[locus] == '?'):
            typings[locus] = newTypings[locus]
        else:
            genotypeAmbiguitySeparator='|'
            alleleCopySeparator='+'
            # In a GLString, '|' is the genotype ambiguity separator.
            # '+' separates copies of a gene at a single locus.
            # Which to use? Most likely, these are two alleles typed separately.
            # But if they are genotypes (Already include a '+' or '|'), we can separate using '|'
            if(genotypeAmbiguitySeparator in typings[locus] or genotypeAmbiguitySeparator in newTypings[locus]
                or alleleCopySeparator in typings[locus] or alleleCopySeparator in newTypings[locus]):
                typings[locus] = typings[locus] + genotypeAmbiguitySeparator + newTypings[locus]
            else:
                typings[locus] = typings[locus] + alleleCopySeparator + newTypings[locus]
    return typings

def constructTypings(allUploads=None, hla=None, token=None, url=None, projectIDs=None, bucket=None, sampleID=None):
    #print('getting typingData for (' + str(hla) + ')')

    fileResults = IhiwRestAccess.getUploadFileNamesByPartialKeyword(uploadTypeFilter=['HML'],
        token=token, url=url,
        fileNameQueries=[str(hla).strip()],
        projectIDs=projectIDs,
        allUploads=allUploads)

    typings={}
    typings['A'] = '?'
    typings['B'] = '?'
    typings['C'] = '?'
    typings['DRB1'] = '?'
    typings['DRB3'] = '?'
    typings['DRB4'] = '?'
    typings['DRB5'] = '?'
    typings['DQB1'] = '?'
    typings['DQA1'] = '?'
    typings['DPB1'] = '?'
    typings['DPA1'] = '?'

    #print('FileResults:' + str(fileResults))

    if(len(fileResults)>0):
        # If it's an HML File Fetch the file, parse for GL Strings, Get the Typing
        #print('I found ' + str(len(fileResults)) + ' files, parsing for GLStrings...')
        for fileResult in fileResults:
            #print('Checking file for glstrings:' + str(fileResult['fileName']))
            currentGlStrings = ParseXml.getGlStringsFromHml(hmlFileName=fileResult['fileName'], s3=s3, bucket=bucket)
            #print('Found this currentGlStrings:' + str(currentGlStrings))
            if currentGlStrings is not None:
                for hmlSampleId in currentGlStrings.keys():
                    if (sampleID is None or sampleID.strip()=='?' or len(str(sampleID).strip()) < 1):
                        # If there is no SampleID Given in DataMatrix, use all the GLStrings.
                        # This will probably result in multiple samples from one HML assigned to this person.
                        typings = updateTypings(typings=typings, newTypings=parseGlStrings(glstrings=currentGlStrings[hmlSampleId]))
                    else:
                        # Check if the data matrix sample id is in the data matrix
                        if(str(sampleID).strip().upper() in str(hmlSampleId).strip().upper()):
                            typings = updateTypings(typings=typings, newTypings=parseGlStrings(glstrings=currentGlStrings[hmlSampleId]))
                        else:
                            # TODO: What do we do if there is no matching sample? We found an HML file but no Sample?
                            # This is another sample in the same HML file.
                            pass

            else:
                print('Warning, No GL Strings found for HML file ' + fileResult['fileName'])

    else:
        typings.update(parseGlStrings(glstrings=[hla]))

    #print('returning typings' + str(typings))
    return typings

def reduceGenotypings(typings=None):
    # Reduce the typings to two fields. Try to maintain a bit of the ambiguity if possible.
    # Keep in mind the GLStrings
    #print('Reducing Genotypings:' + str(typings))
    reducedTyping={}
    for locus in typings.keys():
        #print('Locus = ' + str(locus))
        fullGenotypes = typings[locus]

        if(fullGenotypes is None):
            print('Warning, full genotypes is None for locus ' + str(locus))
            reducedTyping[locus] = typings[locus]
        elif(type(fullGenotypes)==str):
            if(fullGenotypes=='?'):
                # This is fine. Unknown typing.
                reducedTyping[locus]=typings[locus]
            else:
                # Split by genotype ambiguities
                genotypeOptions = set()
                for genotypeString in (fullGenotypes.split('|')):
                    #print('Found genotype option:' + str(genotypeString))
                    # How many alleles does this genotype have? 1 or 2 is normal
                    alleleStrings=genotypeString.split('+')

                    if(len(alleleStrings) in [1,2]):
                        bothAlleles = []
                        for alleleString in alleleStrings:
                            #print('Allele String:' + str(alleleString))
                            alleleOptions = set()
                            for allele in alleleString.split('/'):
                                #print('Allele:' + str(allele))
                                # Get the first 2 fields
                                nomenclatureTokens = allele.split(':')
                                if(len(nomenclatureTokens)==1):
                                    alleleOptions.add(allele)
                                else:
                                    shortAlleleName = nomenclatureTokens[0] + ':' + nomenclatureTokens[1]
                                    expressionCharacter = allele[-1]
                                    if not str.isdigit(expressionCharacter):
                                        #print('This might be a nullallele!:' + allele)
                                        # Check if there are > 2 fields before adding charcter, otherwise we get doubled-up expression characters.
                                        if(len(nomenclatureTokens)>2):
                                            shortAlleleName = shortAlleleName + expressionCharacter
                                    alleleOptions.add(shortAlleleName)
                                #print('alleleOptions:' + str(alleleOptions))
                            bothAlleles.append('/'.join(sorted(list(alleleOptions))))
                            #print('BothAlleles:' + str(bothAlleles))

                        genotypeOptions.add('+'.join(sorted(list(bothAlleles))))

                    else:
                        print('Warning, Apparently there are too many alleles (' + str(len(alleleStrings)) +  ') in a single genotype ' + str(genotypeString))
                reducedTyping[locus] = '|'.join(sorted(list(genotypeOptions)))
        else:
            print('Warning, full genotypes is not a string! for locus ' + str(locus))
            reducedTyping[locus] = typings[locus]

    #print('Reduced Genotypes:' + str(reducedTyping))
    return reducedTyping

def getFullHamlFileNames(token=None, url=None, projectIDs=None, allUploads=None, cellData=None,uploadUser=None):
    hamlFileNames = []

    # Split by comma.
    hamlRawFilenames = cellData.split(',')
    hamlRawFilenames = [s.strip() for s in hamlRawFilenames]

    hamlUploadFilenames = []

    #print('cellData:' + str(cellData))

    for hamlRawFilename in hamlRawFilenames:
        fileResults = IhiwRestAccess.getUploadFileNamesByPartialKeyword(uploadTypeFilter=['HAML','ANTIBODY_CSV'],
            token=token, url=url,
            fileNameQueries=hamlRawFilename,
            projectIDs=projectIDs,
            allUploads=allUploads,
            uploadUser=uploadUser)

        #print('Keyword:' + str(hamlRawFilename))
        #print('fileResults: ' + str([fileResult['fileName'] for fileResult in fileResults]))

        for fileResult in fileResults:
            if fileResult['type'] == 'HAML':
                hamlUploadFilenames.append(fileResult['fileName'])
            elif(fileResult['type'] == 'ANTIBODY_CSV'):
                # TODO: can I just filter the "allUploads" instead of running another rest query here?
                childUploads = IhiwRestAccess.getUploadsByParentId(token=token,url=url,parentId=fileResult['id'])

                #print('childUploads of file ' + str(fileResult['fileName']))

                for childUpload in childUploads:
                    if childUpload['type'] == 'HAML':
                        hamlUploadFilenames.append(childUpload['fileName'])
                    else:
                        pass
            else:
                raise Exception('I found a file that is neither HAML or ANTIBODY_CSV, that is strange:' + str(fileResult))

    # Unique filenames
    hamlUploadFilenames = sorted(list(set(hamlUploadFilenames)))

    if len(hamlUploadFilenames) != 1:
        pass
        # This is common enough we don't need to warn about it..
        #print('Warning: for user ' + str(uploadUser) + ' I found ' + str(len(hamlUploadFilenames)) + ' files of type ' + str('HAML') + ' for the keywords ' + str(cellData))
        #print('\n'.join(hamlUploadFilenames))


    return hamlUploadFilenames

def createAlleleSpecificReport(antibodiesLookup=None, recipientGenotypingsLookup=None, donorGenotypingsLookup=None, bucket=None, reportName=None, isImmunogenic=None):
    print('Creating Allele-Specific antibody report ' + str(reportName))
    #print('recipientGenotypingsLookup:' + str(recipientGenotypingsLookup))

    # A list of alleles
    classISpecificities=set()
    classIISpecificities=set()
    # Positive and negative controls for each patient.
    controlLookup = {}

    # First sort through the antibody data and pull out the stuff we need.
    for transplantationId in antibodiesLookup.keys():
        controlLookup[transplantationId] = {}
        controlLookup[transplantationId]['I'] = {}
        controlLookup[transplantationId]['II'] = {}

        for panel in antibodiesLookup[transplantationId]:
            hlaClass=None
            for specificity in sorted(antibodiesLookup[transplantationId][panel]):
                # Is this panel ClassI or ClassII?
                if(str(specificity).startswith('A*')
                    or str(specificity).startswith('B*')
                    or str(specificity).startswith('C*')
                    or str(specificity).startswith('Cw*')):
                    hlaClass = 'I'
                    #controlLookup[transplantationId][hlaClass]['panel']=panel
                    if panel not in controlLookup[transplantationId][hlaClass].keys():
                        controlLookup[transplantationId][hlaClass][panel] = {}
                        controlLookup[transplantationId][hlaClass][panel]['PC'] = '?'
                        controlLookup[transplantationId][hlaClass][panel]['NC'] = '?'
                    classISpecificities.add(specificity)
                elif(str(specificity).startswith('D')):
                    hlaClass = 'II'
                    #controlLookup[transplantationId][hlaClass]['panel'] = panel
                    if panel not in controlLookup[transplantationId][hlaClass].keys():
                        controlLookup[transplantationId][hlaClass][panel] = {}
                        controlLookup[transplantationId][hlaClass][panel]['PC'] = '?'
                        controlLookup[transplantationId][hlaClass][panel]['NC'] = '?'
                    classIISpecificities.add(specificity)
                elif(specificity.startswith('NC : ')):
                    if (hlaClass in('I','II')):
                        #controlLookup[transplantationId][hlaClass]['NC']=antibodiesLookup[transplantationId][panel][specificity]
                        controlLookup[transplantationId][hlaClass][panel]['NC'] = antibodiesLookup[transplantationId][panel][specificity]
                    else:
                        raise Exception('Is this class I or II?:' + str(specificity))
                elif (specificity.startswith('PC : ')):
                    if (hlaClass in('I','II')):
                        #controlLookup[transplantationId][hlaClass]['PC']=antibodiesLookup[transplantationId][panel][specificity]
                        controlLookup[transplantationId][hlaClass][panel]['PC'] = antibodiesLookup[transplantationId][panel][specificity]
                    else:
                        raise Exception('Is this class I or II?:' + str(specificity))
                else:
                    print('Something went wrong in allele-specific report, HLA class could not be determined:' + str(specificity)
                        + ' transplantationId:' + str(transplantationId)
                        + ' panel:' + str(panel))
                    raise Exception ('Something went wrong in allele-specific report, HLA class could not be determined:' + str(specificity))

    classISpecificities = sorted(list(set(classISpecificities)))
    classIISpecificities = sorted(list(set(classIISpecificities)))


    alleleSpecificReportWorkbook = Workbook()
    alleleSpecificReportWorksheet = alleleSpecificReportWorkbook.active
    alleleSpecificReportWorksheet.freeze_panes = 'B2'

    if(isImmunogenic):
        headers = ['transplantation_id', 'R_A', 'R_B', 'R_C', 'R_DRB1', 'R_DRB3', 'R_DRB4', 'R_DRB5', 'R_DQB1', 'R_DQA1', 'R_DPB1', 'R_DPA1'
            , 'D_A', 'D_B', 'D_C', 'D_DRB1', 'D_DRB3', 'D_DRB4', 'D_DRB5', 'D_DQB1', 'D_DQA1', 'D_DPB1', 'D_DPA1', 'classI-PanelID','classI-PC', 'classI-NC']
    else:
        headers = ['transplantation_id', 'A', 'B', 'C', 'DRB1', 'DRB3', 'DRB4', 'DRB5', 'DQB1', 'DQA1', 'DPB1', 'DPA1', 'classI-PanelID','classI-PC', 'classI-NC']
    headers.extend(classISpecificities)
    headers.extend(['classII-PanelID','classII-PC', 'classII-NC'])
    headers.extend(classIISpecificities)

    #print('Headers:' + str(headers))
    for headerIndex, header in enumerate(headers):
        columnLetter = get_column_letter(headerIndex+1)
        cellIndex = columnLetter + '1'
        alleleSpecificReportWorksheet[cellIndex] = header
        alleleSpecificReportWorksheet.column_dimensions[columnLetter].width = 30

    for transplantationId in antibodiesLookup.keys():

        # Handle the case of multiple panels for a single locus
        classIPanelString = ''
        classIPcString = ''
        classINcString = ''
        for panel in controlLookup[transplantationId]['I'].keys():
            if(len(classIPanelString) > 0):
                # This is not the first panel
                classIPanelString = classIPanelString + ' ; ' + panel
                classIPcString = classIPcString + ' ; ' + controlLookup[transplantationId]['I'][panel]['PC']
                classINcString = classINcString + ' ; ' + controlLookup[transplantationId]['I'][panel]['NC']
            else:
                # This the first or only panel
                classIPanelString = panel
                classIPcString = controlLookup[transplantationId]['I'][panel]['PC']
                classINcString = controlLookup[transplantationId]['I'][panel]['NC']

        classIIPanelString = ''
        classIIPcString = ''
        classIINcString = ''
        for panel in controlLookup[transplantationId]['II'].keys():
            if (len(classIIPanelString) > 0):
                # This is not the first panel
                classIIPanelString = classIIPanelString + ' ; ' + panel
                classIIPcString = classIIPcString + ' ; ' + controlLookup[transplantationId]['II'][panel]['PC']
                classIINcString = classIINcString + ' ; ' + controlLookup[transplantationId]['II'][panel]['NC']
            else:
                # This the first or only panel
                classIIPanelString = panel
                classIIPcString = controlLookup[transplantationId]['II'][panel]['PC']
                classIINcString = controlLookup[transplantationId]['II'][panel]['NC']


        patientData = [str(transplantationId)
            , recipientGenotypingsLookup[transplantationId]['A']
            , recipientGenotypingsLookup[transplantationId]['B']
            , recipientGenotypingsLookup[transplantationId]['C']
            , recipientGenotypingsLookup[transplantationId]['DRB1']
            , recipientGenotypingsLookup[transplantationId]['DRB3']
            , recipientGenotypingsLookup[transplantationId]['DRB4']
            , recipientGenotypingsLookup[transplantationId]['DRB5']
            , recipientGenotypingsLookup[transplantationId]['DQB1']
            , recipientGenotypingsLookup[transplantationId]['DQA1']
            , recipientGenotypingsLookup[transplantationId]['DPB1']
            , recipientGenotypingsLookup[transplantationId]['DPA1']
        ]

        if isImmunogenic:
            patientData.extend([donorGenotypingsLookup[transplantationId]['A']
                , donorGenotypingsLookup[transplantationId]['B']
                , donorGenotypingsLookup[transplantationId]['C']
                , donorGenotypingsLookup[transplantationId]['DRB1']
                , donorGenotypingsLookup[transplantationId]['DRB3']
                , donorGenotypingsLookup[transplantationId]['DRB4']
                , donorGenotypingsLookup[transplantationId]['DRB5']
                , donorGenotypingsLookup[transplantationId]['DQB1']
                , donorGenotypingsLookup[transplantationId]['DQA1']
                , donorGenotypingsLookup[transplantationId]['DPB1']
                , donorGenotypingsLookup[transplantationId]['DPA1']])

        patientData.extend([classIPanelString
            , classIPcString
            , classINcString])

        for classISpecificity in classISpecificities:
            classIMfi = '?'
            for panel in controlLookup[transplantationId]['I'].keys():
                try:
                    classIMfi = antibodiesLookup[transplantationId][panel][classISpecificity]
                except KeyError:
                    # The MFI for this locus is in a different panel.
                    pass
            patientData.append(classIMfi)

        patientData.extend([classIIPanelString
            , classIIPcString
            , classIINcString])

        for classIISpecificity in classIISpecificities:
            classIIMfi = '?'
            for panel in controlLookup[transplantationId]['II'].keys():
                try:
                    classIIMfi = antibodiesLookup[transplantationId][panel][classIISpecificity]
                except KeyError:
                    # The MFI for this locus is in a different panel.
                    pass
            patientData.append(classIIMfi)


        alleleSpecificReportWorksheet.append(patientData)

    outputWorkbookbyteStream = ParseExcel.createBytestreamExcelOutputFile(workbookObject=alleleSpecificReportWorkbook)
    S3_Access.writeFileToS3(newFileName=reportName, bucket=bucket, s3ObjectBytestream=outputWorkbookbyteStream)

def createGlStringFromTypings(sampleTypings=None):
    glString=''

    for locus in ['A','B','C','DRB1','DRB3','DRB4','DRB5','DQB1','DQA1','DPB1','DPA1']:
        if(sampleTypings[locus]!='?'):
            glString = glString + sampleTypings[locus] + '^'

    trimGlString=glString[0:-1]
    return trimGlString

def createImmunogenicEpitopesReport(bucket=None, projectIDs=None, url=None, token=None):
    print('Creating an Immunogenic Epitopes Submission Report for project ids ' + str(projectIDs))

    if url is None:
        url = IhiwRestAccess.getUrl()
        token = IhiwRestAccess.getToken(url=url)

    if(projectIDs is None):
        return
    elif(not isinstance(projectIDs, list)):
        projectIDs = [projectIDs]

    # Convert to String for consistency..
    projectIDs = [str(projectID) for projectID in projectIDs]
    projectString = str('_'.join(projectIDs))

    # preload an upload list to use repeatedly later
    allUploads = IhiwRestAccess.getFilteredUploads(token=token, url=url, projectIDs=projectIDs)
    dataMatrixUploadList = getDataMatrixUploads(projectIDs=projectIDs, token=token, url=url, uploadList=allUploads)

    # This report is the copy of all data in the data matrix, including validation feedback.
    dataMatrixReportWorkbook = Workbook()
    dataMatrixReportWorksheet = dataMatrixReportWorkbook.active
    dataMatrixReportWorksheet.freeze_panes = 'A2'

    # Data Matrix Report Headers
    dataMatrixReportHeaders = ['transplantation_index','data_matrix_filename', 'data_matrix_row_num', 'submitting_user'
        , 'submitting_lab', 'submission_date', 'donor_glstring', 'recipient_glstring', 'transplantation_report']
    dataMatrixHeaders = ImmunogenicEpitopesValidator.getColumnNames(isImmunogenic=True)
    dataMatrixReportHeaders.extend(dataMatrixHeaders)

    for headerIndex, header in enumerate(dataMatrixReportHeaders):
        dataMatrixColumnLetter = get_column_letter(headerIndex+1)
        dataMatrixCellIndex = dataMatrixColumnLetter + '1'
        dataMatrixReportWorksheet[dataMatrixCellIndex] = header
        dataMatrixReportWorksheet.column_dimensions[dataMatrixColumnLetter].width = 35

    # Summary of HLA Genotypes that have been submitted
    summaryWithTypingWorkbook = Workbook()
    summaryWithTypingWorksheet = summaryWithTypingWorkbook.active
    summaryWithTypingWorksheet.freeze_panes = 'A2'

    summaryWithTypingHeaders = ('transplant_id', 'upload_filename', 'row#', 'submitter'
        , 'recipient_sample_id','recipient_hla', 'R_A', 'R_B', 'R_C', 'R_DRB1', 'R_DRB3', 'R_DRB4', 'R_DRB5', 'R_DQB1', 'R_DQA1', 'R_DPB1', 'R_DPA1'
        , 'donor_sample_id', 'donor_hla', 'D_A', 'D_B', 'D_C', 'D_DRB1', 'D_DRB3', 'D_DRB4', 'D_DRB5', 'D_DQB1', 'D_DQA1', 'D_DPB1', 'D_DPA1')

    for headerIndex, header in enumerate(summaryWithTypingHeaders):
        dataMatrixColumnLetter = get_column_letter(headerIndex+1)
        dataMatrixCellIndex = dataMatrixColumnLetter + '1'
        summaryWithTypingWorksheet[dataMatrixCellIndex] = header
        summaryWithTypingWorksheet.column_dimensions[dataMatrixColumnLetter].width = 35

    supportingSpreadsheets = {}

    reportLineIndex = 0

    # I want the first transplantation to be index 0,
    transplantationIndex = -1
    antibodiesPreTxLookup = {}
    antibodiesPostTxLookup = {}
    recipientGenotypingsLookup = {}
    donorGenotypingsLookup = {}


    # Combine data matrices together for summary worksheet..
    for dataMatrixIndex, dataMatrixUpload in enumerate(dataMatrixUploadList):
        print('Checking Validation of this file:' + dataMatrixUpload['fileName'] + ' (' + str(dataMatrixIndex+1) + '/' + str(len(dataMatrixUploadList)) + ') for projects ' + str(projectIDs))
        #print('This is the upload: ' + str(dataMatrixUpload))

        excelFileObject = s3.get_object(Bucket=bucket, Key=dataMatrixUpload['fileName'])
        inputExcelBytes = io.BytesIO(excelFileObject["Body"].read())
        # validateEpitopesDataMatrix returns all the information we need.
        (validationResults, validatedWorkbook) = ImmunogenicEpitopesValidator.validateEpitopesDataMatrix(
            excelFile=inputExcelBytes, isImmunogenic=True, projectIDs=projectIDs, uploadList=allUploads)
        if (validatedWorkbook is not None):
            supportingSpreadsheets[dataMatrixUpload['fileName']] = ParseExcel.createBytestreamExcelOutputFile(
                workbookObject=validatedWorkbook)
            dataMatrixFileName = dataMatrixUpload['fileName']
            submittingUser = dataMatrixUpload['createdBy']['user']['firstName'] + ' ' + \
                             dataMatrixUpload['createdBy']['user']['lastName'] + ':\n' + \
                             dataMatrixUpload['createdBy']['user']['email']
            submittingLab = dataMatrixUpload['createdBy']['lab']['department'] + ', ' + \
                            dataMatrixUpload['createdBy']['lab']['institution']
            submissionDate = dataMatrixUpload['createdAt']

            uploadUserId = dataMatrixUpload['createdBy']['id']
            #print('Found an upload user ID:' + str(uploadUserId))

            # Loop input Workbook data
            # for dataLineIndex, dataLine in enumerate(inputExcelFileData):
            firstSheet = validatedWorkbook[validatedWorkbook.sheetnames[0]]
            for dataLineIndex, dataLine in enumerate(firstSheet.iter_rows(min_row=2)):

                currentExcelRow = dataLineIndex + 2

                recipientSampleId = getDataMatrixValue(columnName='recipient_sample_id', validatedWorkbook=validatedWorkbook, currentExcelRow=currentExcelRow, firstSheet=firstSheet)
                recipientHla = getDataMatrixValue(columnName='recipient_hla', validatedWorkbook=validatedWorkbook, currentExcelRow=currentExcelRow, firstSheet=firstSheet)
                donorSampleId = getDataMatrixValue(columnName='donor_sample_id', validatedWorkbook=validatedWorkbook, currentExcelRow=currentExcelRow, firstSheet=firstSheet)
                donorHla = getDataMatrixValue(columnName='donor_hla', validatedWorkbook=validatedWorkbook, currentExcelRow=currentExcelRow, firstSheet=firstSheet)

                # Check if there is some data here, for now it's good enough if there is some HLA typing included
                if(len(str(recipientHla).strip()) > 1 and len(str(donorHla).strip()) > 1):
                    reportLineIndex += 1
                    transplantationIndex += 1

                    # Get the Donor GLString/HML File
                    # Get the Recipient GLString/HML File
                    recipientTypings = constructTypings(allUploads=allUploads, hla=recipientHla, token=token, url=url, projectIDs=projectIDs, bucket=bucket, sampleID=recipientSampleId)
                    donorTypings = constructTypings(allUploads=allUploads, hla=donorHla, token=token, url=url, projectIDs=projectIDs, bucket=bucket, sampleID=donorSampleId)

                    recipientTypingsSimplified = reduceGenotypings(typings=recipientTypings)
                    donorTypingsSimplified = reduceGenotypings(typings=donorTypings)

                    # Put the typing in the spreadsheets.
                    recipientGenotypingsLookup[transplantationIndex] = recipientTypingsSimplified
                    donorGenotypingsLookup[transplantationIndex] = donorTypingsSimplified

                    # Write data for summaryWithTypingWorksheet
                    summaryWithTypingDataTuple = (str(transplantationIndex), dataMatrixFileName, currentExcelRow, (submittingUser + ', ' + submittingLab)
                        , recipientSampleId ,recipientHla, recipientTypingsSimplified['A'], recipientTypingsSimplified['B']
                        , recipientTypingsSimplified['C'], recipientTypingsSimplified['DRB1'], recipientTypingsSimplified['DRB3'], recipientTypingsSimplified['DRB4'], recipientTypingsSimplified['DRB5']
                        , recipientTypingsSimplified['DQB1'], recipientTypingsSimplified['DQA1'], recipientTypingsSimplified['DPB1'], recipientTypingsSimplified['DPA1']
                        , donorSampleId, donorHla, donorTypingsSimplified['A'], donorTypingsSimplified['B']
                        , donorTypingsSimplified['C'], donorTypingsSimplified['DRB1'], donorTypingsSimplified['DRB3'], donorTypingsSimplified['DRB4'], donorTypingsSimplified['DRB5']
                        , donorTypingsSimplified['DQB1'], donorTypingsSimplified['DQA1'], donorTypingsSimplified['DPB1'], donorTypingsSimplified['DPA1']
                    )
                    summaryWithTypingWorksheet.append(summaryWithTypingDataTuple)

                    # Get the antibody filenames
                    hamlPreTxCellData = getDataMatrixValue(columnName='recipient_haml_pre_tx', validatedWorkbook=validatedWorkbook, currentExcelRow=currentExcelRow, firstSheet=firstSheet)
                    recipHamlPreTxFilenames = getFullHamlFileNames(token=token,url=url, projectIDs=projectIDs, allUploads=allUploads, cellData=hamlPreTxCellData, uploadUser=uploadUserId)

                    hamlPostTxCellData = getDataMatrixValue(columnName='recipient_haml_post_tx', validatedWorkbook=validatedWorkbook, currentExcelRow=currentExcelRow, firstSheet=firstSheet)
                    recipHamlPostTxFilenames = getFullHamlFileNames(token=token,url=url, projectIDs=projectIDs, allUploads=allUploads, cellData=hamlPostTxCellData, uploadUser=uploadUserId)

                    # Create the Antibody Typing Report.
                    if(len(donorTypingsSimplified) > 0 and len(recipientTypingsSimplified) > 0):

                        transplantationReportFileName = 'AntibodyReport_' + dataMatrixUpload['fileName'] + '_Row' + str(
                            currentExcelRow) + '.xlsx'
                        transplantationReportText, preTxAntibodies, postTxAntibodies = getTransplantationReportSpreadsheet(
                            donorTyping=donorTypingsSimplified, recipientTyping=recipientTypingsSimplified,
                            recipHamlPreTxFilenames=recipHamlPreTxFilenames, recipHamlPostTxFilenames=recipHamlPostTxFilenames,
                            s3=s3, bucket=bucket, transplantationIndex=transplantationIndex, recipientSampleId=recipientSampleId)
                        supportingSpreadsheets[transplantationReportFileName] = transplantationReportText

                        #print('The antibody report returned these values:' + str(preTxAntibodies) + str(postTxAntibodies))

                        antibodiesPreTxLookup[transplantationIndex] = preTxAntibodies
                        antibodiesPostTxLookup[transplantationIndex] = postTxAntibodies



                    # Copy the columns (including colors and notes) into the new spreadsheet

                    dataMatrixReportLine = [str(transplantationIndex), dataMatrixUpload['fileName'], str(currentExcelRow),submittingUser,submittingLab,submissionDate
                        ,createGlStringFromTypings(recipientTypingsSimplified), createGlStringFromTypings(donorTypingsSimplified),transplantationReportFileName]

                    # Pull the data from the data matrix and add it to the report worksheet.
                    columnNameLookup = validatedWorkbook.columnNameLookup
                    for headerIndex, header in enumerate(dataMatrixHeaders):
                        if(header in columnNameLookup.keys()):
                            dataMatrixColumnLetter = validatedWorkbook.columnNameLookup[header]
                            dataMatrixCellIndex = dataMatrixColumnLetter + str(currentExcelRow)
                            dataMatrixCell = firstSheet[dataMatrixCellIndex]
                            dataMatrixReportLine.append(str(dataMatrixCell.value))
                        else:
                            dataMatrixReportLine.append('?')

                    dataMatrixReportWorksheet.append(dataMatrixReportLine)

                    # Apply Red background color and comments from the data matrix
                    # This is done in a separate loop, it seems to mess up the line indexing otherwise.
                    for headerIndex, header in enumerate(dataMatrixHeaders):
                        if (header in columnNameLookup.keys()):
                            dataMatrixColumnLetter = validatedWorkbook.columnNameLookup[header]
                            dataMatrixCellIndex = dataMatrixColumnLetter + str(currentExcelRow)
                            dataMatrixCell = firstSheet[dataMatrixCellIndex]
                            dataMatrixCellColor = dataMatrixCell.fill.fgColor.value
                            reportCellIndex = str(get_column_letter(headerIndex + 10)) + str(reportLineIndex + 1)

                            try:
                                if (dataMatrixCellColor == '00FF0000'):
                                    # This means there was an error in this cell!
                                    dataMatrixReportWorksheet[reportCellIndex].fill = PatternFill("solid", fgColor=dataMatrixCellColor)
                                    dataMatrixReportWorksheet[reportCellIndex].comment = Comment(dataMatrixCell.comment.text, 'Data Matrix Validator')
                            except Exception as e:
                                print('That did not work:' + str(e))

                        else:
                            pass

                    # TODO: Size the columns and rows better

                    # TODO: Text wrapping?



                else:
                    #print('Empty Data line.')
                    pass


        else:
            print('No workbook data was found for data matrix ' + str(dataMatrixUpload['fileName']))
            print('Upload ID of missing data matrix:' + str(dataMatrixUpload['id']))


    dataMatrixReportFileName = 'Project.' + projectString+ '.DataMatrixReport.xlsx'
    outputWorkbookbyteStream = ParseExcel.createBytestreamExcelOutputFile(workbookObject=dataMatrixReportWorkbook)
    S3_Access.writeFileToS3(newFileName=dataMatrixReportFileName, bucket=bucket, s3ObjectBytestream=outputWorkbookbyteStream)

    summaryWithTypingFileName = 'Project.' + projectString+ '.SampleSummary.xlsx'
    outputWorkbookbyteStream = ParseExcel.createBytestreamExcelOutputFile(workbookObject=summaryWithTypingWorkbook)
    S3_Access.writeFileToS3(newFileName=summaryWithTypingFileName, bucket=bucket, s3ObjectBytestream=outputWorkbookbyteStream)

    preTxAlleleSpecificReportName = 'Project.' + str('_'.join(projectIDs)) + '.AlleleSpecificPreTx.xlsx'
    createAlleleSpecificReport(antibodiesLookup=antibodiesPreTxLookup, recipientGenotypingsLookup=recipientGenotypingsLookup
        , donorGenotypingsLookup=donorGenotypingsLookup, bucket=bucket, reportName=preTxAlleleSpecificReportName)

    postTxAlleleSpecificReportName = 'Project.' + str('_'.join(projectIDs)) + '.AlleleSpecificPostTx.xlsx'
    createAlleleSpecificReport(antibodiesLookup=antibodiesPostTxLookup, recipientGenotypingsLookup=recipientGenotypingsLookup
        , donorGenotypingsLookup=donorGenotypingsLookup, bucket=bucket, reportName=postTxAlleleSpecificReportName)


    # create zip file
    zipFileName = 'Project.' + str('_'.join(projectIDs)) + '.TransplantationReports.zip'
    zipFileStream = io.BytesIO()
    supportingFileZip = zipfile.ZipFile(zipFileStream, 'a', zipfile.ZIP_DEFLATED, False)

    for transplantationReportFileName in supportingSpreadsheets.keys():
        #print('Adding file ' + str(transplantationReportFileName) + ' to ' + str(zipFileName))
        supportingFileZip.writestr(transplantationReportFileName, supportingSpreadsheets[transplantationReportFileName])

    supportingFileZip.close()
    S3_Access.writeFileToS3(newFileName=zipFileName, bucket=bucket, s3ObjectBytestream=zipFileStream)

def createNonImmunogenicEpitopesReport(bucket=None, projectIDs=None, url=None, token=None):
    print('Creating an non-Immunogenic Epitopes Submission Report for project ids ' + str(projectIDs))

    if url is None:
        url = IhiwRestAccess.getUrl()
        token = IhiwRestAccess.getToken(url=url)

    if(projectIDs is None):
        return
    elif(not isinstance(projectIDs, list)):
        projectIDs = [projectIDs]

    # Convert to String for consistency..
    projectIDs = [str(projectID) for projectID in projectIDs]
    projectString = str('_'.join(projectIDs))

    # preload an upload list to use repeatedly later
    allUploads = IhiwRestAccess.getUploadsByProjects(token=token, url=url, projectIDs=projectIDs)
    dataMatrixUploadList = getDataMatrixUploads(projectIDs=projectIDs, token=token, url=url, uploadList=allUploads)

    # This report is the copy of all data in the data matrix, including validation feedback.
    dataMatrixReportWorkbook = Workbook()
    dataMatrixReportWorksheet = dataMatrixReportWorkbook.active
    dataMatrixReportWorksheet.freeze_panes = 'A2'

    # Data Matrix Report Headers
    dataMatrixReportHeaders = ['datarow_id','data_matrix_filename', 'data_matrix_row_num', 'submitting_user'
        , 'submitting_lab', 'submission_date', 'recipient_glstring']
    dataMatrixHeaders = ImmunogenicEpitopesValidator.getColumnNames(isImmunogenic=False)
    dataMatrixReportHeaders.extend(dataMatrixHeaders)

    for headerIndex, header in enumerate(dataMatrixReportHeaders):
        dataMatrixColumnLetter = get_column_letter(headerIndex+1)
        dataMatrixCellIndex = dataMatrixColumnLetter + '1'
        dataMatrixReportWorksheet[dataMatrixCellIndex] = header
        dataMatrixReportWorksheet.column_dimensions[dataMatrixColumnLetter].width = 35

    # Summary of HLA Genotypes that have been submitted
    summaryWithTypingWorkbook = Workbook()
    summaryWithTypingWorksheet = summaryWithTypingWorkbook.active
    summaryWithTypingWorksheet.freeze_panes = 'A2'

    summaryWithTypingHeaders = ('datarow_id', 'upload_filename', 'row#', 'submitter'
        , 'sample_id','recipient_hla', 'A', 'B', 'C', 'DRB1', 'DRB3', 'DRB4', 'DRB5', 'DQB1', 'DQA1', 'DPB1', 'DPA1')

    for headerIndex, header in enumerate(summaryWithTypingHeaders):
        dataMatrixColumnLetter = get_column_letter(headerIndex+1)
        dataMatrixCellIndex = dataMatrixColumnLetter + '1'
        summaryWithTypingWorksheet[dataMatrixCellIndex] = header
        summaryWithTypingWorksheet.column_dimensions[dataMatrixColumnLetter].width = 35

    supportingSpreadsheets = {}
    reportLineIndex = 0

    # I want the first data row to be index 0,
    dataRowIndexIndex = -1
    recipientAntibodiesLookup = {}
    recipientGenotypingsLookup = {}

    # Combine data matrices together for summary worksheet..
    for dataMatrixIndex, dataMatrixUpload in enumerate(dataMatrixUploadList):
        print('Checking Validation of this file:' + dataMatrixUpload['fileName'] + ' (' + str(dataMatrixIndex+1) + '/' + str(len(dataMatrixUploadList)) + ') for projects ' + str(projectIDs))

        excelFileObject = s3.get_object(Bucket=bucket, Key=dataMatrixUpload['fileName'])
        inputExcelBytes = io.BytesIO(excelFileObject["Body"].read())
        # validateEpitopesDataMatrix returns all the information we need.
        (validationResults, validatedWorkbook) = ImmunogenicEpitopesValidator.validateEpitopesDataMatrix(
            excelFile=inputExcelBytes, isImmunogenic=False, projectIDs=projectIDs, uploadList=allUploads)

        if (validatedWorkbook is not None):
            supportingSpreadsheets[dataMatrixUpload['fileName']] = ParseExcel.createBytestreamExcelOutputFile(
                workbookObject=validatedWorkbook)
            dataMatrixFileName = dataMatrixUpload['fileName']
            submittingUser = dataMatrixUpload['createdBy']['user']['firstName'] + ' ' + \
                             dataMatrixUpload['createdBy']['user']['lastName'] + ':\n' + \
                             dataMatrixUpload['createdBy']['user']['email']
            submittingLab = dataMatrixUpload['createdBy']['lab']['department'] + ', ' + \
                            dataMatrixUpload['createdBy']['lab']['institution']
            submissionDate = dataMatrixUpload['createdAt']

            uploadUserId = dataMatrixUpload['createdBy']['id']
            print('Found an upload user ID:' + str(uploadUserId))

            # Loop input Workbook data
            # for dataLineIndex, dataLine in enumerate(inputExcelFileData):
            firstSheet = validatedWorkbook[validatedWorkbook.sheetnames[0]]
            for dataLineIndex, dataLine in enumerate(firstSheet.iter_rows(min_row=2)):

                currentExcelRow = dataLineIndex + 2

                recipientSampleId = getDataMatrixValue(columnName='recipient_sample_id', validatedWorkbook=validatedWorkbook, currentExcelRow=currentExcelRow, firstSheet=firstSheet)
                recipientHla = getDataMatrixValue(columnName='recipient_hla', validatedWorkbook=validatedWorkbook, currentExcelRow=currentExcelRow, firstSheet=firstSheet)

                # Check if there is some data here, for now it's good enough if there is some HLA typing included
                if(len(str(recipientHla).strip()) > 1):
                    reportLineIndex += 1
                    dataRowIndexIndex += 1

                    # Get the Donor GLString/HML File
                    # Get the Recipient GLString/HML File
                    recipientTypings = constructTypings(allUploads=allUploads, hla=recipientHla, token=token, url=url, projectIDs=projectIDs, bucket=bucket, sampleID=recipientSampleId)
                    recipientTypingsSimplified = reduceGenotypings(typings=recipientTypings)

                    # Put the typing in the spreadsheets.
                    recipientGenotypingsLookup[dataRowIndexIndex] = recipientTypingsSimplified

                    # Write data for summaryWithTypingWorksheet
                    summaryWithTypingDataTuple = (str(dataRowIndexIndex), dataMatrixFileName, currentExcelRow, (submittingUser + ', ' + submittingLab)
                        , recipientSampleId ,recipientHla, recipientTypingsSimplified['A'], recipientTypingsSimplified['B']
                        , recipientTypingsSimplified['C'], recipientTypingsSimplified['DRB1'], recipientTypingsSimplified['DRB3'], recipientTypingsSimplified['DRB4'], recipientTypingsSimplified['DRB5']
                        , recipientTypingsSimplified['DQB1'], recipientTypingsSimplified['DQA1'], recipientTypingsSimplified['DPB1'], recipientTypingsSimplified['DPA1']
                    )
                    summaryWithTypingWorksheet.append(summaryWithTypingDataTuple)

                    # Get the antibody filenames
                    hamlCellData = getDataMatrixValue(columnName='recipient_haml', validatedWorkbook=validatedWorkbook, currentExcelRow=currentExcelRow, firstSheet=firstSheet)
                    recipHamlFilenames = getFullHamlFileNames(token=token, url=url, projectIDs=projectIDs, allUploads=allUploads, cellData=hamlCellData, uploadUser=uploadUserId)


                    # Create the Antibody Typing Report.
                    if(len(recipientTypingsSimplified) > 0):

                        transplantationReportFileName = 'AntibodyReport_' + dataMatrixUpload['fileName'] + '_Row' + str(currentExcelRow) + '.xlsx'
                        transplantationReportText, preTxAntibodies, postTxAntibodies = getTransplantationReportSpreadsheet(
                            donorTyping=recipientTypingsSimplified, recipientTyping=recipientTypingsSimplified,
                            recipHamlPreTxFilenames=recipHamlFilenames, recipHamlPostTxFilenames=recipHamlFilenames,
                            s3=s3, bucket=bucket, transplantationIndex=dataRowIndexIndex, recipientSampleId=recipientSampleId)
                        supportingSpreadsheets[transplantationReportFileName] = transplantationReportText

                        #print('The antibody report returned these values:' + str(preTxAntibodies) + str(postTxAntibodies))

                        recipientAntibodiesLookup[dataRowIndexIndex] = preTxAntibodies



                    # Copy the columns (including colors and notes) into the new spreadsheet
                    dataMatrixReportLine = [str(dataRowIndexIndex), dataMatrixUpload['fileName'], str(currentExcelRow),submittingUser,submittingLab,submissionDate
                        ,createGlStringFromTypings(recipientTypingsSimplified)]

                    # Pull the data from the data matrix and add it to the report worksheet.
                    columnNameLookup = validatedWorkbook.columnNameLookup
                    for headerIndex, header in enumerate(dataMatrixHeaders):
                        if(header in columnNameLookup.keys()):
                            dataMatrixColumnLetter = validatedWorkbook.columnNameLookup[header]
                            dataMatrixCellIndex = dataMatrixColumnLetter + str(currentExcelRow)
                            dataMatrixCell = firstSheet[dataMatrixCellIndex]
                            dataMatrixReportLine.append(str(dataMatrixCell.value))
                        else:
                            dataMatrixReportLine.append('?')

                    dataMatrixReportWorksheet.append(dataMatrixReportLine)

                    # Apply Red background color and comments from the data matrix
                    # This is done in a separate loop, it seems to mess up the line indexing otherwise.
                    for headerIndex, header in enumerate(dataMatrixHeaders):
                        if (header in columnNameLookup.keys()):
                            dataMatrixColumnLetter = validatedWorkbook.columnNameLookup[header]
                            dataMatrixCellIndex = dataMatrixColumnLetter + str(currentExcelRow)
                            dataMatrixCell = firstSheet[dataMatrixCellIndex]
                            dataMatrixCellColor = dataMatrixCell.fill.fgColor.value
                            reportCellIndex = str(get_column_letter(headerIndex + 8)) + str(reportLineIndex + 1)

                            try:
                                if (dataMatrixCellColor == '00FF0000'):
                                    # This means there was an error in this cell!
                                    dataMatrixReportWorksheet[reportCellIndex].fill = PatternFill("solid", fgColor=dataMatrixCellColor)
                                    dataMatrixReportWorksheet[reportCellIndex].comment = Comment(dataMatrixCell.comment.text, 'Data Matrix Validator')
                            except Exception as e:
                                print('That did not work:' + str(e))

                        else:
                            pass

                    # TODO: Size the columns and rows better

                    # TODO: Text wrapping?



                else:
                    #print('Empty Data line.')
                    pass


        else:
            print('No workbook data was found for data matrix ' + str(dataMatrixUpload['fileName']))
            print('Upload ID of missing data matrix:' + str(dataMatrixUpload['id']))


    dataMatrixReportFileName = 'Project.' + projectString+ '.DataMatrixReport.xlsx'
    outputWorkbookbyteStream = ParseExcel.createBytestreamExcelOutputFile(workbookObject=dataMatrixReportWorkbook)
    S3_Access.writeFileToS3(newFileName=dataMatrixReportFileName, bucket=bucket, s3ObjectBytestream=outputWorkbookbyteStream)

    summaryWithTypingFileName = 'Project.' + projectString+ '.SampleSummary.xlsx'
    outputWorkbookbyteStream = ParseExcel.createBytestreamExcelOutputFile(workbookObject=summaryWithTypingWorkbook)
    S3_Access.writeFileToS3(newFileName=summaryWithTypingFileName, bucket=bucket, s3ObjectBytestream=outputWorkbookbyteStream)

    preTxAlleleSpecificReportName = 'Project.' + str('_'.join(projectIDs)) + '.AlleleSpecific.xlsx'
    createAlleleSpecificReport(antibodiesLookup=recipientAntibodiesLookup, recipientGenotypingsLookup=recipientGenotypingsLookup
       , bucket=bucket, reportName=preTxAlleleSpecificReportName, isImmunogenic=False)

    '''
    postTxAlleleSpecificReportName = 'Project.' + str('_'.join(projectIDs)) + '.AlleleSpecificPostTx.xlsx'
    createAlleleSpecificReport(antibodiesLookup=antibodiesPostTxLookup, recipientGenotypingsLookup=recipientGenotypingsLookup
        , donorGenotypingsLookup=donorGenotypingsLookup, bucket=bucket, reportName=postTxAlleleSpecificReportName)
    
 
        
    # create zip file
    zipFileName = 'Project.' + str('_'.join(projectIDs)) + '.TransplantationReports.zip'
    zipFileStream = io.BytesIO()
    supportingFileZip = zipfile.ZipFile(zipFileStream, 'a', zipfile.ZIP_DEFLATED, False)

    for transplantationReportFileName in supportingSpreadsheets.keys():
        #print('Adding file ' + str(transplantationReportFileName) + ' to ' + str(zipFileName))
        supportingFileZip.writestr(transplantationReportFileName, supportingSpreadsheets[transplantationReportFileName])

    supportingFileZip.close()
    S3_Access.writeFileToS3(newFileName=zipFileName, bucket=bucket, s3ObjectBytestream=zipFileStream)
    '''

def getDataMatrixUploads(projectIDs=None, token=None, url=None, uploadList=None):
    # collect all data matrix files.
    if(uploadList is None):
        uploadList = IhiwRestAccess.getUploads(token=token, url=url)
    #print('I found these uploads:' + str(uploadList))
    #print('This is my upload list:' + str(uploadList))
    print('Parsing ' + str(len(uploadList)) + ' uploads to find data matrices for project(s) ' + str(projectIDs) + '..')
    if(not isinstance(projectIDs, list)):
        projectIDs = [projectIDs]
    dataMatrixUploadList = []
    for upload in uploadList:
        uploadId = str(upload['project']['id'])
        #print('UploadID:' + uploadId)
        #print('ProjectIDs:' + str(projectIDs))
        if (uploadId in projectIDs):
            if (upload['type'] == 'PROJECT_DATA_MATRIX'):
                dataMatrixUploadList.append(upload)
            else:
                #print('Disregarding this upload because it is not a data matrix.')
                pass
        else:
            #print('Disregarding this upload because it is not in our project.')
            pass
    print(
        'I found a total of ' + str(len(dataMatrixUploadList)) + ' data matrices for project' + str(projectIDs) + '.\n')
    return dataMatrixUploadList







