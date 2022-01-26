from boto3 import client
s3 = client('s3')
from sys import exc_info
import json
import urllib
from time import sleep
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from io import BytesIO

# For importing common methods, may be in the same directory when deployed as a package
# TODO: Fix these imports. I think I can remove the try/catch here by changing the project root directory.
try:
    import IhiwRestAccess
    import Validation
    import S3_Access
    import ParseExcel
except Exception as e:
    print('Failed in importing files: ' + str(e))
    from Common import IhiwRestAccess
    from Common import Validation
    from Common import S3_Access
    from Common import ParseExcel

def immunogenic_epitope_handler(event, context):
    print('I found the immunogenic epitope validation handler.')
    # TODO: This should handle the case

    # This is the AWS Lambda handler function.
    uploadDetails = {}
    uploadDetails['is_valid'] = False
    uploadDetails['validation_feedback'] = 'The input event payload does not seem to contain the expected upload data. Something went wrong in the step function flowchart.'
    uploadDetails['validator_type'] = 'DATA_MATRIX'

    try:
        print('Inside Try')
        print('Event info, this should have [Input][Payload]:' + str(event))
        # Sleep 1 second, enough time to make sure the file is available.
        #sleep(1)
        # Fetching the upload detail bundle from the step function payload
        if "Input" in event and "Payload" in event['Input']:
            uploadDetails = event['Input']['Payload']
        else:
            print(str(uploadDetails['validation_feedback']))
            return uploadDetails

        bucket = uploadDetails['bucket']
        # print('bucket:' + str(bucket))

        excelKey = uploadDetails['file_name']
        print('Excel Filename:' + excelKey)

        validationResults = None
        # Is this an excel file? It should have the xlsx extension.

        url = uploadDetails['url']
        token = uploadDetails['token']


        projectName = str(uploadDetails['project_name'])
        projectID =  str(uploadDetails['project_id'])
        print('The upload is for this project:' + str(projectName) + ',id=' + str(projectID))
        fileType = uploadDetails['upload_type']
        print('this upload is file type:' + str(fileType))


        # Is this a data Matrix?
        if (fileType == 'PROJECT_DATA_MATRIX'):
            immunogenicEpitopeProjectNumber=str(IhiwRestAccess.getProjectID(projectName='immunogenic_epitopes'))
            nonImmunogenicEpitopeProjectNumber=str(IhiwRestAccess.getProjectID(projectName='non_immunogenic_epitopes'))
            dqImmunogenicityProjectNumber=str(IhiwRestAccess.getProjectID(projectName='dq_immunogenicity'))

            if (projectID == immunogenicEpitopeProjectNumber or projectID == dqImmunogenicityProjectNumber):
                print('This is the Immunogenic Epitopes project!')
                validatorType='IMMUNOGENIC_EPITOPES'
                isImmunogenic = True
            elif (projectID == nonImmunogenicEpitopeProjectNumber):
                print('This is the Non Immunogenic Epitopes project!')
                validatorType = 'NON_IMMUNOGENIC_EPITOPES'
                isImmunogenic = False
            else:
                print('This is not the (Non) Immunogenic Epitopes project! I will not validate it. Double-check the project names')
                return None

            excelFileObject = s3.get_object(Bucket=bucket, Key=excelKey)
            inputExcelBytes = BytesIO(excelFileObject["Body"].read())

            (validationResults, outputReportWorkbook) = validateEpitopesDataMatrix(excelFile=inputExcelBytes, isImmunogenic=isImmunogenic, projectIDs=projectID, url=url, token=token)
            if(len(validationResults) == 0):
                validationFeedback='Valid'
            else:
                validationFeedback = ''
                for validationError in validationResults:
                    # for validationColumnName in validationErrorRow.keys():
                    # currentValidationResult = validationErrorRow[validationColumnName]
                    if (validationError is not None and len(validationError) > 0):
                        validationFeedback = validationFeedback + validationError + ';\n'

            print('validation results were retrieved, attempting to set validation status.')
            print('ValidationResults:(\n' + str(validationResults) + '\n)')
            uploadDetails['is_valid'] = len(validationResults) == 0
            uploadDetails['validation_feedback'] = validationFeedback
            uploadDetails['validator_type'] = validatorType

            createValidationReport(isReportValid=len(validationResults) == 0, parentUploadFileName=excelKey
                , outputReportWorkbook=outputReportWorkbook, bucket=bucket, token=token, url=url, validatorType=validatorType)

        elif (fileType == 'XLSX'):
            print('This file is the excel report file. I do not need to validate it (again)')
            # Returning 'None' So we don't overwrite the existing validation status.
            # TODO: It might be better to detect a validation status and pass it to the next step. This is better for step functions,
            #   But to do that I need to access the parent Project_Data_Matrix and get it's validation status

        else:
            print('the file type ' + str(fileType) + ' is not a project data matrix, i will not validate it.')

        return uploadDetails

    except Exception as e:
        print('Exception:\n' + str(e) + '\n' + str(exc_info()))
        return str(e)


def validateCell(columnIndexLookup=None, currentCell=None, uploadList=None, hmlUploadList=None, hamlUploadList=None, isImmunogenic=None):
    columnLetter = get_column_letter(currentCell.column)
    headerName = columnIndexLookup[columnLetter]
    cellValue = currentCell.value
    #print(str(currentCell) + ' contains data for column ' + str(columnLetter) + ' which is header ' + headerName)

    currentValidationFeedback = ''
    if(isImmunogenic):
        if(headerName=='recipient_hla'): currentValidationFeedback = Validation.validateHlaGenotypeEntry(query=cellValue, searchList=hmlUploadList, allowPartialMatch=True, columnName=headerName,  uploadList=uploadList)
        elif(headerName == 'recipient_sample_id'): currentValidationFeedback = ''
        elif(headerName == 'donor_sample_id'): currentValidationFeedback = ''
        elif(headerName=='recipient_haml_pre_tx'): currentValidationFeedback = Validation.validateUniqueEntryInList(query=cellValue, searchList=hamlUploadList, allowPartialMatch=True, columnName=headerName)
        elif(headerName=='recipient_haml_post_tx'): currentValidationFeedback = Validation.validateUniqueEntryInList(query=cellValue, searchList=hamlUploadList, allowPartialMatch=True, columnName=headerName)
        elif(headerName=='recipient_sex'): currentValidationFeedback = Validation.validateMaleFemale(query=cellValue, columnName=headerName)
        elif(headerName=='recipient_year_of_birth'): currentValidationFeedback = Validation.validateNumber(query=cellValue, columnName=headerName)
        elif(headerName=='recipient_pregnancies'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName)
        elif(headerName=='recipient_transfusions'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='recipient_dialysis_date'): currentValidationFeedback = Validation.validateDate(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='recipient_deceased_date'): currentValidationFeedback = Validation.validateDate(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='donor_year_of_birth'): currentValidationFeedback = Validation.validateNumber(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='recipient_blood_group'): currentValidationFeedback = Validation.validateBloodGroup(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='donor_source_type'): currentValidationFeedback = Validation.validateDonorSourceType(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='donor_hla'): currentValidationFeedback = Validation.validateHlaGenotypeEntry(query=cellValue, searchList=hmlUploadList, allowPartialMatch=True, columnName=headerName, uploadList=uploadList)
        elif(headerName=='donor_sex'): currentValidationFeedback = Validation.validateMaleFemale(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='donor_blood_group'): currentValidationFeedback = Validation.validateBloodGroup(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='transplantation_date'): currentValidationFeedback = Validation.validateDate(query=cellValue, columnName=headerName)
        elif(headerName=='transplant_organ_category'): currentValidationFeedback = Validation.validateOrganCategory(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='prozone_pre_tx'): currentValidationFeedback = Validation.validateProzoneType(query=cellValue, columnName=headerName)
        elif(headerName=='prozone_post_tx'): currentValidationFeedback = Validation.validateProzoneType(query=cellValue, columnName=headerName)
        elif(headerName=='availability_pre_tx'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='availability_post_tx'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='date_antibody_pre_tx'): currentValidationFeedback = Validation.validateDate(query=cellValue, columnName=headerName)
        elif(headerName=='date_antibody_post_tx'): currentValidationFeedback = Validation.validateDate(query=cellValue, columnName=headerName)
        elif(headerName=='immune_suppr_post_tx'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='organ_status_post_tx'): currentValidationFeedback = Validation.validateOrganStatus(query=cellValue, columnName=headerName, required=False)
        else: currentValidationFeedback = 'Unknown Column Name:' + str(headerName)
    else:
        if(headerName=='recipient_hla'): currentValidationFeedback = Validation.validateHlaGenotypeEntry(query=cellValue, searchList=hmlUploadList, allowPartialMatch=True, columnName=headerName, uploadList=uploadList)
        elif (headerName == 'recipient_sample_id'): currentValidationFeedback = ''
        elif(headerName=='recipient_haml'): currentValidationFeedback = Validation.validateUniqueEntryInList(query=cellValue, searchList=hamlUploadList, allowPartialMatch=True, columnName=headerName)
        elif(headerName=='prozone'): currentValidationFeedback = Validation.validateProzoneType(query=cellValue, columnName=headerName)
        elif(headerName=='sample_availability'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName)
        elif(headerName=='age_recipient_years'): currentValidationFeedback = Validation.validateNumber(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='recipient_sex'): currentValidationFeedback = Validation.validateMaleFemale(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='recipient_pregnancies'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='recipient_transfusions'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName, required=False)
        elif(headerName=='recipient_transplantations'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName, required=False)
        else: currentValidationFeedback = 'Unknown Column Name:' + str(headerName)

    #print('Validation Feedback:' + str(currentValidationFeedback))
    # Set Cell background color and feedback text
    if(currentValidationFeedback is not None and len(currentValidationFeedback)>0):
        errorColor = 'FF0000'  # Red
        currentCell.fill = PatternFill("solid", fgColor=errorColor)
        currentCell.comment=Comment(currentValidationFeedback, 'Data Matrix Validator')

    return currentValidationFeedback

def validateEpitopesDataMatrix(excelFile=None, isImmunogenic=None, projectIDs=None, url=None, token=None, uploadList=None):
    # This method returns a tuple, with (list of validation results, Report Excel Excel Data)
    validationErrors = []
    inputExcelData = None
    if (isImmunogenic == None):
        print(
            'Please pass isImmunogenic=True/False, to specify whether we should validatate Immunogenic or NonImmunogenic epitopes.')
        return ('Cannot determine if this is a immunogenic or non immunogenic matrix. Please pass isImmunogenic=True/False', inputExcelData, validationErrors)
    elif (isImmunogenic):
        print('Validating Immunogenic Epitopes.')
        epitopeColumnNames = getColumnNames(isImmunogenic=True)
    else:
        print('Validating Non Immunogenic Epitopes.')
        epitopeColumnNames = getColumnNames(isImmunogenic=False)

    try:
        xlWorkbook = load_workbook(filename=excelFile)
    except Exception:
        print('Error: Could not load data matrix!')
        return (['Error: Could not load data matrix!'],None)

    inputColumns = {}
    firstSheet = xlWorkbook[xlWorkbook.sheetnames[0]]
    for colIndexRaw, col in enumerate(firstSheet.iter_cols()):
        if(col[0].value is not None and len(str(col[0].value))>0):
            headerName = str(col[0].value).strip().lower()
            columnLetter = get_column_letter(col[0].column)
            inputColumns[headerName]=columnLetter

    missingColumns=[column for column in epitopeColumnNames if column not in inputColumns]
    extraColumns=[column for column in inputColumns if column not in epitopeColumnNames]

    if(len(missingColumns) > 0):
        print('Missing Columns!:' + str(missingColumns))
        validationErrors.append('Missing Columns!:' + str(missingColumns))
    if (len(extraColumns) > 0):
        print('Extra Columns!:' + str(extraColumns))
        validationErrors.append('Extra Columns!:' + str(extraColumns))

    if(url is None):
        url = IhiwRestAccess.getUrl()
    if(token is None):
        token = IhiwRestAccess.getToken(url=url)
    if(token is None or url is None or len(url)<1 or len(token) < 1):
        return('Could not aquire a login token.', inputExcelData, validationErrors)
    try:
        if(uploadList is None):
            uploadList = IhiwRestAccess.getUploadsByProjects(token=token, url=url, projectIDs=projectIDs)
        hmlUploadList = Validation.createFileListFromUploads(uploads=uploadList, projectFilter=projectIDs, fileTypeFilter='HML')
        hamlUploadList = Validation.createFileListFromUploads(uploads=uploadList, projectFilter=projectIDs, fileTypeFilter='HAML')
        antibodyCsvUploadList = Validation.createFileListFromUploads(uploads=uploadList, projectFilter=projectIDs, fileTypeFilter='ANTIBODY_CSV')
        # Add the antibody_CSV files to the haml file list.
        # TODO: I'm currently ignoring the Antibody CSV files. Is that okay?
        #hamlUploadList.extend(antibodyCsvUploadList)

    except Exception as e:
        print('Exception when getting list of uploads:\n' + str(e) + '\n' + str(exc_info()))
        return ('Exception when getting list of uploads:\n' + str(e) , inputExcelData, validationErrors)

    # Invert the columns to make a lookup
    columnIndexLookup = {v: k for k, v in inputColumns.items()}
    # Do more specific validation of the columns. Check that each column is valid.
    # This is project-specific so there are different validations for each column.
    #for dataRowIndex, dataRow in enumerate(inputExcelData):
    for rowIndexRaw, row in enumerate(firstSheet.iter_rows(min_row=2)):
        for cell in row:
            if(cell.value is not None and len(str(cell.value))>0):
                currentValidationFeedback = validateCell(columnIndexLookup=columnIndexLookup, currentCell=cell, uploadList=uploadList, hmlUploadList=hmlUploadList, hamlUploadList=hamlUploadList, isImmunogenic=isImmunogenic)
                if(currentValidationFeedback is not None and len(currentValidationFeedback)>0):
                    validationErrors.append(currentValidationFeedback)

    reportWorkbook=xlWorkbook
    reportWorkbook.columnNameLookup = inputColumns # Sneakily embedding a column lookup in the result worksheet.
    return validationErrors, reportWorkbook

def getColumnNames(isImmunogenic=True):
    if (isImmunogenic):
        return [
            'recipient_sample_id'
            , 'recipient_hla'
            , 'recipient_haml_pre_tx'
            , 'recipient_haml_post_tx'
            , 'recipient_sex'
            , 'recipient_year_of_birth'
            , 'recipient_pregnancies'
            , 'recipient_transfusions'
            , 'recipient_dialysis_date'
            , 'recipient_deceased_date'
            , 'donor_year_of_birth'
            , 'recipient_blood_group'
            , 'donor_sample_id'
            , 'donor_source_type'
            , 'donor_hla'
            , 'donor_sex'
            , 'donor_blood_group'
            , 'transplantation_date'
            , 'transplant_organ_category'
            , 'prozone_pre_tx'
            , 'prozone_post_tx'
            , 'availability_pre_tx'
            , 'availability_post_tx'
            , 'date_antibody_pre_tx'
            , 'date_antibody_post_tx'
            , 'immune_suppr_post_tx'
            , 'organ_status_post_tx'
        ]
    else:
        return [
            'recipient_sample_id'
            , 'recipient_hla'
            , 'recipient_haml'
            , 'prozone'
            , 'sample_availability'
            , 'age_recipient_years'
            , 'recipient_sex'
            , 'recipient_pregnancies'
            , 'recipient_transfusions'
            , 'recipient_transplantations'

        ]

def createValidationReport(isReportValid=None, parentUploadFileName=None, outputReportWorkbook=None, bucket=None, token=None, url=None, validatorType=None):
    print('Creating a validation Report.')
    reportFileName = parentUploadFileName + '.Validation_Report.xlsx'
    reportStreamData = ParseExcel.createBytestreamExcelOutputFile(workbookObject=outputReportWorkbook)

    # if it already exists, we don't need to create the upload object for the report.
    # This throws a HTTP Error 404 if the file doesn't exist
    # I think it's not great to use try/catch for program logic, maybe there's a better solution
    existingUpload = IhiwRestAccess.getUploadIfExists(token=token, url=url, fileName=reportFileName)
    if(existingUpload is None):
        print('There is no child upload for the report file ' + str(reportFileName) + ' so I will create one.')
        IhiwRestAccess.createConvertedUploadObject(newUploadFileName=reportFileName, newUploadFileType='XLSX', token=token, url=url,previousUploadFileName=parentUploadFileName)
    else:
        print('There is already a child upload(' + str(existingUpload) + ') for report file ' + str(reportFileName) + ' so I will not create one.')

    # TODO: It might be better to detect a validation status in the handler for the XLSX Report, and pass it to the next step.
    #   This is better for step functions, and lets the handler handle the xlsx files individually.
    #   But to do that I need to access the parent Project_Data_Matrix and get it's validation status
    IhiwRestAccess.setValidationStatus(uploadFileName=reportFileName, isValid=isReportValid,
        validationFeedback='Download this report for Detailed Validation Results.'
        , validatorType=validatorType + '_REPORT', token=token, url=url)

    S3_Access.writeFileToS3(newFileName=reportFileName, bucket=bucket, s3ObjectBytestream=reportStreamData)