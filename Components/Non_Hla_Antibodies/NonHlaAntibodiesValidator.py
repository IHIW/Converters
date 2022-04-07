from boto3 import client
s3 = client('s3')
from sys import exc_info
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from io import BytesIO

# For importing common methods, may be in the same directory when deployed as a package
# TODO: Fix these imports. I think I can remove the try/catch here by changing the project root directory.
try:
    import IhiwRestAccess
    import Validation
    import S3_Access
    import ParseExcel
except Exception as e:
    print('ImmunogenicEpitopesValidator Failed in importing files: ' + str(e))
    from Common import IhiwRestAccess
    from Common import Validation
    from Common import S3_Access
    from Common import ParseExcel

def non_hla_antibodies_handler(event, context):
    print('I found the immunogenic epitope validation handler.')
    # TODO: This should handle the case

    # This is the AWS Lambda handler function.
    uploadDetails = {}
    uploadDetails['is_valid'] = False
    uploadDetails['validation_feedback'] = 'The input event payload does not seem to contain the expected upload data. Something went wrong in the step function flowchart.'
    uploadDetails['validator_type'] = 'DATA_MATRIX'

    try:
        print('Inside Try')
        print('Event info, this should have [Input][Payload]:' + str(event))
        # Sleep 1 second, enough time to make sure the file is available.
        #sleep(1)
        # Fetching the upload detail bundle from the step function payload
        if "Input" in event and "Payload" in event['Input']:
            uploadDetails = event['Input']['Payload']
        else:
            print(str(uploadDetails['validation_feedback']))
            return uploadDetails

        bucket = uploadDetails['bucket']
        # print('bucket:' + str(bucket))

        excelKey = uploadDetails['file_name']
        print('Excel Filename:' + excelKey)
        uploadId = uploadDetails['id']

        validationResults = None
        # Is this an excel file? It should have the xlsx extension.

        url = uploadDetails['url']
        token = uploadDetails['token']

        projectName = str(uploadDetails['project_name'])
        projectID =  str(uploadDetails['project_id'])
        print('The upload is for this project:' + str(projectName) + ',id=' + str(projectID))
        fileType = uploadDetails['upload_type']
        print('this upload is file type:' + str(fileType))


        # Is this a data Matrix?
        if (fileType == 'PROJECT_DATA_MATRIX'):
            immunogenicEpitopeProjectNumber=str(IhiwRestAccess.getProjectID(projectName='non_hla_antibodies'))

            if (projectID == immunogenicEpitopeProjectNumber):
                print('This is the NON_HLA_ANTIBODIES project!')
                validatorType='NON_HLA_ANTIBODIES'
                isImmunogenic = True
            else:
                print('This is not the Non HLA Antibodies! I will not validate it. Double-check the project names')
                return None

            excelFileObject = s3.get_object(Bucket=bucket, Key=excelKey)
            inputExcelBytes = BytesIO(excelFileObject["Body"].read())

            (validationResults, outputReportWorkbook) = validateNonHlaAntibodiesDataMatrix(excelFile=inputExcelBytes, projectIDs=projectID, url=url, token=token)
            if(len(validationResults) == 0):
                validationFeedback='Valid'
            else:
                validationFeedback = ''
                for validationError in validationResults:
                    # for validationColumnName in validationErrorRow.keys():
                    # currentValidationResult = validationErrorRow[validationColumnName]
                    if (validationError is not None and len(validationError) > 0):
                        validationFeedback = validationFeedback + validationError + ';\n'

            print('validation results were retrieved, attempting to set validation status.')
            print('ValidationResults:(\n' + str(validationResults) + '\n)')
            uploadDetails['is_valid'] = len(validationResults) == 0
            uploadDetails['validation_feedback'] = validationFeedback
            uploadDetails['validator_type'] = validatorType

            createValidationReport(isReportValid=len(validationResults) == 0, parentUploadFileName=excelKey, parentId=uploadId
                , outputReportWorkbook=outputReportWorkbook, bucket=bucket, token=token, url=url, validatorType=validatorType)

        elif (fileType == 'XLSX'):
            print('This file is the excel report file. I do not need to validate it (again)')
            # Returning 'None' So we don't overwrite the existing validation status.
            # TODO: It might be better to detect a validation status and pass it to the next step. This is better for step functions,
            #   But to do that I need to access the parent Project_Data_Matrix and get it's validation status

        else:
            print('the file type ' + str(fileType) + ' is not a project data matrix, i will not validate it.')

        return uploadDetails

    except Exception as e:
        print('Exception:\n' + str(e) + '\n' + str(exc_info()))
        return str(e)


def validateCell(columnIndexLookup=None, currentCell=None, uploadList=None, antibodyCsvUploadList=None):
    columnLetter = get_column_letter(currentCell.column)
    headerName = columnIndexLookup[columnLetter]
    cellValue = currentCell.value
    #print(str(currentCell) + ' contains data for column ' + str(columnLetter) + ' which is header ' + headerName)
    currentValidationFeedback = ''

    if(headerName == 'patient_identifier'): currentValidationFeedback = Validation.validateTextExists(query=cellValue, columnName=headerName)
    elif(headerName == 'year_of_transplant'): currentValidationFeedback = Validation.validateNumber(query=cellValue, columnName=headerName, required=True)
    elif(headerName == 'patient_year_of_birth'): currentValidationFeedback = Validation.validateNumber(query=cellValue, columnName=headerName, required=True)
    elif(headerName == 'patient_sex'): currentValidationFeedback = Validation.validateMaleFemale(query=cellValue, columnName=headerName, required=True)
    elif(headerName == 'patient_ethnicity'): currentValidationFeedback = ''
    elif(headerName == 'rejection'): currentValidationFeedback = Validation.validateBoolean(query=cellValue, columnName=headerName, required=True)
    elif(headerName == 'rejection_type'): currentValidationFeedback = Validation.validateRejectionType(query=cellValue, columnName=headerName, required=False)
    elif(headerName == 'graft_number'): currentValidationFeedback = Validation.validateNumber(query=cellValue, columnName=headerName, required=False)
    elif(headerName == 'disease_aetiology'): currentValidationFeedback = Validation.validateDiseaseAetiology(query=cellValue, columnName=headerName, required=True)
    elif(headerName == 'pre_tx_sample_id'): currentValidationFeedback = Validation.validateTextExists(query=cellValue, columnName=headerName)
    elif(headerName == 'pre_tx_sample_date'): currentValidationFeedback = Validation.validateDate(query=cellValue, columnName=headerName, required=True)
    elif(headerName == 'pre_tx_csv_immucor'): currentValidationFeedback = Validation.validateUniqueEntryInList(query=cellValue, searchList=antibodyCsvUploadList, allowPartialMatch=True, columnName=headerName, delimiter=',', required=False)
    elif(headerName == 'pre_tx_csv_onelambda'): currentValidationFeedback = Validation.validateUniqueEntryInList(query=cellValue, searchList=antibodyCsvUploadList, allowPartialMatch=True, columnName=headerName, delimiter=',', required=False)
    elif(headerName == 'post_tx_antibody_timing'): currentValidationFeedback = Validation.validateNumber(query=cellValue, columnName=headerName, required=False)
    elif(headerName == 'post_tx_sample_id'): currentValidationFeedback = ''
    elif(headerName == 'post_tx_sample_date'): currentValidationFeedback = Validation.validateDate(query=cellValue, columnName=headerName, required=False)
    elif(headerName == 'post_tx_csv_immucor'): currentValidationFeedback = Validation.validateUniqueEntryInList(query=cellValue, searchList=antibodyCsvUploadList, allowPartialMatch=True, columnName=headerName, delimiter=',', required=False)
    elif(headerName == 'post_tx_csv_onelambda'): currentValidationFeedback = Validation.validateUniqueEntryInList(query=cellValue, searchList=antibodyCsvUploadList, allowPartialMatch=True, columnName=headerName, delimiter=',', required=False)
    else: currentValidationFeedback = 'Unknown Column Name:' + str(headerName)

    #print('Validation Feedback:' + str(currentValidationFeedback))
    # Set Cell background color and feedback text
    if(currentValidationFeedback is not None and len(currentValidationFeedback)>0):
        errorColor = 'FF0000'  # Red
        currentCell.fill = PatternFill("solid", fgColor=errorColor)
        currentCell.comment=Comment(currentValidationFeedback, 'Data Matrix Validator')

    return currentValidationFeedback

def validateNonHlaAntibodiesDataMatrix(excelFile=None, projectIDs=None, url=None, token=None, uploadList=None):
    # This method returns a tuple, with (list of validation results, Report Excel Excel Data)
    validationErrors = []
    inputExcelData = None

    epitopeColumnNames = getColumnNames()

    try:
        xlWorkbook = load_workbook(filename=excelFile)
    except Exception:
        print('Error: Could not load data matrix!')
        return (['Error: Could not load data matrix!'],None)

    inputColumns = {}
    firstSheet = xlWorkbook[xlWorkbook.sheetnames[0]]
    for colIndexRaw, col in enumerate(firstSheet.iter_cols()):
        if(col[0].value is not None and len(str(col[0].value))>0):
            headerName = str(col[0].value).strip().lower()
            columnLetter = get_column_letter(col[0].column)
            inputColumns[headerName]=columnLetter

    missingColumns=[column for column in epitopeColumnNames if column not in inputColumns]
    extraColumns=[column for column in inputColumns if column not in epitopeColumnNames]

    if(len(missingColumns) > 0):
        print('Missing Columns!:' + str(missingColumns))
        validationErrors.append('Missing Columns!:' + str(missingColumns))
    if (len(extraColumns) > 0):
        print('Extra Columns!:' + str(extraColumns))
        validationErrors.append('Extra Columns!:' + str(extraColumns))

    if(url is None):
        url = IhiwRestAccess.getUrl()
    if(token is None):
        token = IhiwRestAccess.getToken(url=url)
    if(token is None or url is None or len(url)<1 or len(token) < 1):
        return('Could not aquire a login token.', inputExcelData, validationErrors)
    try:
        if(uploadList is None):
            uploadList = IhiwRestAccess.getUploadsByProjects(token=token, url=url, projectIDs=projectIDs)

        antibodyCsvUploadList = Validation.createFileListFromUploads(uploads=uploadList, projectFilter=projectIDs, fileTypeFilter='ANTIBODY_CSV')

    except Exception as e:
        print('Exception when getting list of uploads:\n' + str(e) + '\n' + str(exc_info()))
        return ('Exception when getting list of uploads:\n' + str(e) , inputExcelData, validationErrors)

    # Invert the columns to make a lookup
    columnIndexLookup = {v: k for k, v in inputColumns.items()}
    # Do more specific validation of the columns. Check that each column is valid.
    # This is project-specific so there are different validations for each column.
    #for dataRowIndex, dataRow in enumerate(inputExcelData):
    for rowIndexRaw, row in enumerate(firstSheet.iter_rows(min_row=2)):
        for cell in row:
            if(cell.value is not None and len(str(cell.value))>0):
                currentValidationFeedback = validateCell(columnIndexLookup=columnIndexLookup, currentCell=cell, uploadList=uploadList, antibodyCsvUploadList=antibodyCsvUploadList)
                if(currentValidationFeedback is not None and len(currentValidationFeedback)>0):
                    validationErrors.append(currentValidationFeedback)

    reportWorkbook=xlWorkbook
    reportWorkbook.columnNameLookup = inputColumns # Sneakily embedding a column lookup in the result worksheet.
    return validationErrors, reportWorkbook

def getColumnNames():
    return [
        'patient_identifier'
        , 'year_of_transplant'
        , 'patient_year_of_birth'
        , 'patient_sex'
        , 'patient_ethnicity'
        , 'rejection'
        , 'rejection_type'
        , 'graft_number'
        , 'disease_aetiology'
        , 'pre_tx_sample_id'
        , 'pre_tx_sample_date'
        , 'pre_tx_csv_immucor'
        , 'pre_tx_csv_onelambda'
        , 'post_tx_antibody_timing'
        , 'post_tx_sample_id'
        , 'post_tx_sample_date'
        , 'post_tx_csv_immucor'
        , 'post_tx_csv_onelambda'
    ]

def createValidationReport(isReportValid=None, parentUploadFileName=None, parentId=None, outputReportWorkbook=None, bucket=None, token=None, url=None, validatorType=None):
    print('Creating a validation Report.')
    reportFileName = parentUploadFileName + '.Validation_Report.xlsx'
    reportStreamData = ParseExcel.createBytestreamExcelOutputFile(workbookObject=outputReportWorkbook)

    try:
        # Delete the Children of this parent Upload.
        print('Looking for children of this upload object..')
        childUploads = IhiwRestAccess.getUploadsByParentId(token=token, url=url, parentId=parentId)
        print('Found these children:' + str(childUpload['id']) for childUpload in childUploads)

        for childUpload in childUploads:
            if (childUpload['type'] == 'XLSX'):
                childUploadFileName = childUpload['fileName']
                childUploadId = childUpload['id']
                deleteResponse = IhiwRestAccess.deleteUpload(token=token, url=url, uploadId=childUploadId)
                print('Found this response from deleting the child:' + str(deleteResponse))
    except Exception as e:
        print('Warning! Could not delete the previous child upload:' + str(e))

    try:
        IhiwRestAccess.createConvertedUploadObject(newUploadFileName=reportFileName, newUploadFileType='XLSX', token=token, url=url, previousUploadFileName=parentUploadFileName)
    except Exception as e:
        print('Warning! Could not Create the upload object for the child:' + str(e))

    IhiwRestAccess.setValidationStatus(uploadFileName=reportFileName, isValid=isReportValid,
        validationFeedback='Download this report for Detailed Validation Results.'
        , validatorType=validatorType + '_REPORT', token=token, url=url)

    S3_Access.writeFileToS3(newFileName=reportFileName, bucket=bucket, s3ObjectBytestream=reportStreamData)