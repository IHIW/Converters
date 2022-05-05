from boto3 import client
from openpyxl import Workbook
#from openpyxl.utils import get_column_letter
#from openpyxl.styles import PatternFill
#from openpyxl.comments import Comment

try:
    import IhiwRestAccess
    import ParseExcel
    #import ParseXml
    #import Validation
    import S3_Access
    #import NonHlaAntibodiesValidator
except Exception as e:
    print('Failed in importing files: ' + str(e))
    from Common import IhiwRestAccess
    from Common import ParseExcel
    #from Common import ParseXml
    ##from Common import Validation
    from Common import S3_Access
    import NonHlaAntibodiesValidator

s3 = client('s3')

def createNonHlaAntibodiesReport(bucket=None, projectIDs=None, url=None, token=None):
    print('Creating an Non HLA Antibodies Report for project ids ' + str(projectIDs))

    # TODO: Nothing here yet, take some code from Epitopes and paste together the data matrices?
    #   Add a .zip file for all submitted data or something

    if url is None:
        url = IhiwRestAccess.getUrl()
        token = IhiwRestAccess.getToken(url=url)

    if(projectIDs is None):
        return
    elif(not isinstance(projectIDs, list)):
        projectIDs = [projectIDs]

    # Convert to String for consistency..
    projectIDs = [str(projectID) for projectID in projectIDs]
    projectString = str('_'.join(projectIDs))

    S3_Access.createProjectZipFile(bucket=bucket, url=url, token=token, projectIDs=projectIDs, fileTypeFilter=['OTHER', 'ANTIBODY_CSV','PROJECT_DATA_MATRIX'])


    # preload an upload list to use repeatedly later
    allUploads = IhiwRestAccess.getFilteredUploads(token=token, url=url, projectIDs=projectIDs)
    #dataMatrixUploadList = getDataMatrixUploads(projectIDs=projectIDs, token=token, url=url, uploadList=allUploads)

    # This report is the copy of all data in the data matrix, including validation feedback.
    dataMatrixReportWorkbook = Workbook()
    dataMatrixReportWorksheet = dataMatrixReportWorkbook.active
    dataMatrixReportWorksheet.freeze_panes = 'A2'

    dataMatrixReportFileName = 'Project.' + projectString+ '.DataMatrixReport.xlsx'
    outputWorkbookbyteStream = ParseExcel.createBytestreamExcelOutputFile(workbookObject=dataMatrixReportWorkbook)
    S3_Access.writeFileToS3(newFileName=dataMatrixReportFileName, bucket=bucket, s3ObjectBytestream=outputWorkbookbyteStream)
