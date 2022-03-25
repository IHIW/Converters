from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from tempfile import NamedTemporaryFile

def createBytestreamExcelOutputFile(workbookObject=None):
    with NamedTemporaryFile() as tmp:
        workbookObject.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        return stream

def alleleListFromTypings(typings=None):
    #print('typings=' + str(typings))
    alleleList = []

    for locus in typings.keys():
        modifiedString = typings[locus].replace('^', '|').replace('+', '|').replace('/', '|').replace('HLA-', '')
        alleles = modifiedString.split('|')
        alleleList.extend(alleles)
    alleleList = sorted(list(set(alleleList)))

    # Sometimes '?' is used to represent an untyped locus.
    try:
        alleleList.remove('?')
    except Exception as e:
        pass

    return alleleList




def createExcelTransplantationReport(donorTyping=None, recipientTyping=None, preTxFileNames=['PreTXFileName']
        , postTxFileNames=['PostTXFileName'], recipPreTxAntibodyData=None, recipPostTxAntibodyData=None, transReport=None, reportName=None, transplantationIndex=None):

    if(reportName is None):
        reportName = 'Transplantation Report'

    # These are apparently storing the antibodies that match specifiities of the recipient. Not sure what I planned to do with these but unused now.
    preTxAntibodies={}
    postTxAntibodies={}

    # We can add a tab to existing report or create a new one.
    if(transReport is None):
        transReport = Workbook()
        reportWorksheet = transReport.active
        reportWorksheet.title = reportName
    else:
        reportWorksheet = transReport.create_sheet(reportName)

    donorAlleles = alleleListFromTypings(typings=donorTyping)
    recipAlleles = alleleListFromTypings(typings=recipientTyping)

    donorColor='FFC2B3' # Red
    recipientColor='99FFFF' # Blue
    bothColor='E6CCFF' # Purple
    blankColor=None#'FFFFFF' # White

    reportWorksheet['A1'] = 'Transplantation ID:' + str(transplantationIndex)
    reportWorksheet['A2'] = 'Donor Typing: ' + str(donorTyping)
    #reportWorksheet['B1'] = str(donorTyping)
    reportWorksheet['A2'].fill = PatternFill("solid", fgColor=donorColor)
    reportWorksheet['A3'] = 'Recipient Typing: ' + str(recipientTyping)
    #reportWorksheet['B2'] = str(recipientTyping)
    reportWorksheet['A3'].fill = PatternFill("solid", fgColor=recipientColor)
    reportWorksheet['A5'] = 'PreTX Bead Data'
    reportWorksheet['A6'] = str(preTxFileNames)
    reportWorksheet['D5'] = 'PostTX Bead Data'
    reportWorksheet['D6'] = str(postTxFileNames)

    combinedSpecificities = set()

    for panel in recipPreTxAntibodyData.keys():
        for specificity in recipPreTxAntibodyData[panel].keys():
            combinedSpecificities.add(specificity)
    for panel in recipPostTxAntibodyData.keys():
        for specificity in recipPostTxAntibodyData[panel].keys():
            combinedSpecificities.add(specificity)
    combinedSpecificities = sorted(list(set(combinedSpecificities)))

    currentRow = 6 # Start at 6

    for specificity in combinedSpecificities:

        currentRow += 1

        # Color Cells?
        specificityDonorMatch = typingMatch(alleleList=donorAlleles, queryAllele=specificity)
        specificityRecipientMatch = typingMatch(alleleList=recipAlleles, queryAllele=specificity)

        beadFound = False
        for panel in recipPreTxAntibodyData.keys():
            if specificity in recipPreTxAntibodyData[panel].keys():
                beadFound = True
                reportWorksheet['A' + str(currentRow)] = specificity
                reportWorksheet['B' + str(currentRow)] = str(recipPreTxAntibodyData[panel][specificity])
                if specificityRecipientMatch:
                    preTxAntibodies[specificity] = str(recipPreTxAntibodyData[panel][specificity])
            else:
                pass
        if not beadFound:
            reportWorksheet['A' + str(currentRow)] = specificity
            reportWorksheet['B' + str(currentRow)] = '?'

        beadFound = False
        for panel in recipPostTxAntibodyData.keys():
            if specificity in recipPostTxAntibodyData[panel].keys():
                beadFound = True
                reportWorksheet['D' + str(currentRow)] = specificity
                reportWorksheet['E' + str(currentRow)] = str(recipPostTxAntibodyData[panel][specificity])
                if specificityRecipientMatch:
                    postTxAntibodies[specificity] = str(recipPostTxAntibodyData[panel][specificity])
            else:
                pass
        if not beadFound:
            reportWorksheet['D' + str(currentRow)] = specificity
            reportWorksheet['E' + str(currentRow)] = '?'

        # Color cells.
        if(specificityDonorMatch and specificityRecipientMatch):
            cellColor = bothColor
        elif (specificityDonorMatch and not specificityRecipientMatch):
            cellColor = donorColor
        elif(not specificityDonorMatch and  specificityRecipientMatch):
            cellColor = recipientColor
        else:
            cellColor = blankColor

        if cellColor is not None:
            reportWorksheet['A' + str(currentRow)].fill = PatternFill("solid", fgColor=cellColor)
            #reportWorksheet['B' + str(currentRow)].fill = PatternFill("solid", fgColor=cellColor)
            reportWorksheet['D' + str(currentRow)].fill = PatternFill("solid", fgColor=cellColor)
            #reportWorksheet['E' + str(currentRow)].fill = PatternFill("solid", fgColor=cellColor)


    # Some formatting to make things pretty
    reportWorksheet.column_dimensions['A'].width = 35
    reportWorksheet.column_dimensions['B'].width = 15
    reportWorksheet.column_dimensions['C'].width = 2
    reportWorksheet.column_dimensions['D'].width = 35
    reportWorksheet.column_dimensions['E'].width = 15
    reportWorksheet.merge_cells('A2:Z2')
    reportWorksheet.merge_cells('A3:Z3')
    reportWorksheet.merge_cells('A5:B5')
    reportWorksheet.merge_cells('A6:B6')
    reportWorksheet.merge_cells('D5:E5')
    reportWorksheet.merge_cells('D6:E6')
    reportWorksheet.freeze_panes = 'A7'

    # Return it as a stream, so we can consume it or save it later.
    return createBytestreamExcelOutputFile(workbookObject=transReport)


def typingMatch(alleleList=None, queryAllele=None):
    for allele in alleleList:
        if(allele.strip().upper().replace('HLA-','') in queryAllele.strip().upper().replace('HLA-','') ):
            return True
    return False

